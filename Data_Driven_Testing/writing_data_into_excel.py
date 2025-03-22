import openpyxl


# file = "C:\\POC\\testdata.xlsx"
#
# workbook =openpyxl.load_workbook(file)
# sheet=workbook["Sheet1"]
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="welcome"
# workbook.save(file)

# write different data into the file
file = "C:\\POC\\testdata.xlsx"

workbook =openpyxl.load_workbook(file)
sheet=workbook["data"]
sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "rohini"
sheet.cell(1,3).value = "SDET"
workbook.save(file)
